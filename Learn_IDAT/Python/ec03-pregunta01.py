from openpyxl import Workbook
from openpyxl import load_workbook

wb1 = load_workbook('datos.xlsx') #Cargamos el excel

ws =  wb1['Hoja1'] #Accedemos a la primera hoja

#Se muestra todas las personas que tiene deuda
print("#####Personas con Deuda#####")
for i in range(10):
    if i>=4:
        deud = 'F'+str(i)
        deuda = ws[deud]
        nomb = 'B'+str(i)
        nombre = ws[nomb]
        apel = 'C'+str(i)
        apellido = ws[apel]
        sema = 'E'+str(i)
        semana = ws[sema]
        print(f'Buenas noches Sr(a) {nombre.value} {apellido.value}, le hablamos de IDAT, queremos hacerle recordar que tiene una deuda pendiente con nosotros de: {deuda.value} soles y tiene un retraso de: {semana.value} semanas”')
        