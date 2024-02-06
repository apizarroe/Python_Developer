from openpyxl import Workbook
from openpyxl import load_workbook

wb1 = load_workbook('datos.xlsx') #Cargamos el excel

ws =  wb1['Hoja1'] #Accedemos a la primera hoja
ws['E4'] = 'V' #Modificamos la celda E4
wb1.save('datos.xlsx')

x = ws['C7'] #Extraemos el dato de la celda C7
print(x.value)

#Se muestra todos los nombres de las personas
for i in range(10):
    if i>=4:
        nomb = 'B'+str(i)
        nombre = ws[nomb]
        print(nombre.value)
        
#Se actualiza todos los datos de una columna
for i in range(10):
    if i>=4:
        nomb = 'E'+str(i)
        ws[nomb] = 'X'
        wb1.save('datos.xlsx')
        
#Se muestra todas las personas que tiene deuda
print("#####Personas con Deuda#####")
for i in range(10):
    if i>=4:
        deud = 'F'+str(i)
        deuda = ws[deud]
        if (deuda.value > 0):
            nomb = 'B'+str(i)
            nombre = ws[nomb]
            apel = 'C'+str(i)
            apellido = ws[apel]
            print(nombre.value)
            print(f'Buenas noches estimada/o {nombre.value} {apellido.value}, somos la institución IDAT, por favor cancelar su deuda que es de un total de {deuda.value} soles')
        
