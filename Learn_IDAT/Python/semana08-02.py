from openpyxl import Workbook
from openpyxl import load_workbook
wb1 = load_workbook('nuevo.xlsx') #Cargamos el excel
ws1 = wb1.create_sheet('hoja1') #Creamos una hoja nueva
ws2 = wb1.create_sheet('hoja2') #Creamos una hoja nueva
wb1.save('nuevo.xlsx')
