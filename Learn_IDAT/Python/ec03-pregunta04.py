import serial,time
from datetime import datetime
from openpyxl import Workbook
from openpyxl import load_workbook

i = 2
valor = ""
puerto = serial.Serial('COM3',9600) #Configuramos puerto y velocidad
#time.sleep(2)
wb1 = Workbook() #Crasmos una nueva hoja de Calculo
wb1.save('analogicos.xlsx') #guardamos ('nombre.extension')
ws1 = wb1.create_sheet('potenciometro') #Creamos una hoja nueva
wb1.save('analogicos.xlsx')

while(i < 52):
    valor = puerto.readline().decode('utf-8').rstrip()
    ws =  wb1['potenciometro'] #Accedemos a la primera hoja
    val = 'C'+str(i)
    ws[val] = valor
    hor = 'B'+str(i)
    current_time = datetime.now()
    ws[hor] = current_time.strftime("%H:%M:%S")
    i = i+1

wb1.save('analogicos.xlsx')
puerto.close()
